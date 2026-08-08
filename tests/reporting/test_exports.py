from decimal import Decimal
from io import BytesIO

import pytest
from django.core.exceptions import PermissionDenied
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import load_workbook

from apps.imports.models import SourceFile
from apps.ledger.choices import InvoiceDirection, MoneyDirection
from apps.reconciliation.choices import ReconciliationDirection
from apps.reconciliation.services import AllocationInput, create_reconciliation
from apps.reporting.exports import build_reconciliation_export, sanitize_excel_value
from tests.builders import make_invoice, make_transaction


@pytest.fixture
def reconciliation(finance_user):
    invoice = make_invoice(finance_user, invoice_number="INV-EXPORT")
    transaction = make_transaction(
        finance_user,
        amount=Decimal("1000.00"),
        counterparty=invoice.counterparty,
    )
    return create_reconciliation(
        actor=finance_user,
        direction=ReconciliationDirection.PURCHASE_PAYMENT,
        allocations=[AllocationInput(invoice.id, transaction.id, Decimal("1000.00"))],
    )


def _add_source_file(batch, name):
    content = name.encode()
    return SourceFile.objects.create(
        batch=batch,
        file=SimpleUploadedFile(name, content),
        original_name=name,
        sha256=f"{name:0<64}"[:64],
        size=len(content),
    )


@pytest.mark.django_db
def test_reconciliation_export_contains_source_links(
    finance_user, reconciliation, settings, tmp_path
):
    settings.MEDIA_ROOT = tmp_path
    allocation = reconciliation.allocations.first()
    invoice_source = _add_source_file(allocation.invoice.import_batch, "invoice-source")
    transaction_source = _add_source_file(
        allocation.transaction.import_batch, "transaction-source"
    )
    workbook = build_reconciliation_export([reconciliation.id], actor=finance_user)
    sheet = workbook["核销明细"]

    assert sheet["A2"].value == str(reconciliation.id)
    assert sheet["H2"].value == allocation.invoice.invoice_number
    assert sheet["E2"].number_format == "0.00"
    assert sheet["G1"].value == "资金来源批次ID"
    assert sheet["K1"].value == "资金来源文件ID"
    assert sheet["L1"].value == "发票来源批次ID"
    assert sheet["M1"].value == "发票来源文件ID"
    assert sheet["G2"].value == str(allocation.transaction.import_batch_id)
    assert sheet["K2"].value == str(transaction_source.id)
    assert sheet["L2"].value == str(allocation.invoice.import_batch_id)
    assert sheet["M2"].value == str(invoice_source.id)
    assert sheet.freeze_panes == "A2"


@pytest.mark.django_db
def test_reconciliation_export_saves_aware_datetimes_as_local_naive_values(
    finance_user, reconciliation
):
    workbook = build_reconciliation_export([reconciliation.id], actor=finance_user)
    content = BytesIO()

    workbook.save(content)
    content.seek(0)
    reloaded = load_workbook(content, data_only=True)
    sheet = reloaded["核销明细"]

    assert sheet["B2"].value.tzinfo is None
    assert sheet["D2"].value.tzinfo is None
    assert sheet["D2"].value.hour == 17


@pytest.mark.django_db
def test_reconciliation_export_creates_all_finance_sheets(finance_user, reconciliation):
    workbook = build_reconciliation_export([reconciliation.id], actor=finance_user)

    assert set(workbook.sheetnames) == {"应收", "应付", "未匹配资金", "核销明细", "导入异常"}
    assert all(workbook[name].freeze_panes == "A2" for name in workbook.sheetnames)


@pytest.mark.django_db
def test_finance_sheets_label_their_respective_source_links(finance_user):
    receivable = make_invoice(finance_user, direction=InvoiceDirection.OUTPUT)
    payable = make_invoice(finance_user)
    transaction = make_transaction(finance_user, direction=MoneyDirection.INFLOW)

    workbook = build_reconciliation_export([], actor=finance_user)

    assert workbook["应收"]["G1"].value == "发票来源批次ID"
    assert workbook["应付"]["H1"].value == "发票来源文件ID"
    assert workbook["未匹配资金"]["G1"].value == "资金来源批次ID"
    assert workbook["未匹配资金"]["H1"].value == "资金来源文件ID"
    assert workbook["导入异常"]["G1"].value == "来源批次ID"
    assert workbook["导入异常"]["H1"].value == "来源文件ID"
    assert workbook["应收"]["G2"].value == str(receivable.import_batch_id)
    assert workbook["应付"]["G2"].value == str(payable.import_batch_id)
    assert workbook["未匹配资金"]["G2"].value == str(transaction.import_batch_id)


@pytest.mark.django_db
def test_reconciliation_export_rejects_owner(owner_user, reconciliation):
    with pytest.raises(PermissionDenied, match="仅财务"):
        build_reconciliation_export([reconciliation.id], actor=owner_user)


@pytest.mark.django_db
def test_reconciliation_export_escapes_untrusted_database_strings(finance_user):
    invoice = make_invoice(finance_user, invoice_number="=IMPORTXML(A1)")
    invoice.counterparty.name = "@external-party"
    invoice.counterparty.save(update_fields=["name"])
    transaction = make_transaction(
        finance_user,
        amount=Decimal("1000.00"),
        counterparty=invoice.counterparty,
    )
    reconciliation = create_reconciliation(
        actor=finance_user,
        direction=ReconciliationDirection.PURCHASE_PAYMENT,
        allocations=[AllocationInput(invoice.id, transaction.id, Decimal("1000.00"))],
    )

    sheet = build_reconciliation_export([reconciliation.id], actor=finance_user)["核销明细"]

    assert sheet["C2"].value == "'@external-party"
    assert sheet["H2"].value == "'=IMPORTXML(A1)"


@pytest.mark.django_db
def test_unmatched_funds_export_masks_misconfigured_full_account_identifier(finance_user):
    transaction = make_transaction(finance_user, direction=MoneyDirection.INFLOW)
    transaction.account.masked_identifier = "121902307610001"
    transaction.account.save(update_fields=["masked_identifier"])

    sheet = build_reconciliation_export([], actor=finance_user)["未匹配资金"]

    assert sheet["F2"].value == "***********0001"


def test_excel_export_escapes_formula_strings_without_coercing_numbers_or_dates():
    assert sanitize_excel_value("=1+1") == "'=1+1"
    assert sanitize_excel_value("+SUM(A1:A2)") == "'+SUM(A1:A2)"
    assert sanitize_excel_value("-danger") == "'-danger"
    assert sanitize_excel_value("@danger") == "'@danger"
    assert sanitize_excel_value(Decimal("10.00")) == Decimal("10.00")
