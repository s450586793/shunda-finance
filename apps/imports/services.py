import csv
import io
import json
from dataclasses import asdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import xlrd
from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.db import IntegrityError, transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.roles import require_finance_actor
from apps.core.audit import record_audit
from apps.core.uploads import validate_upload_signature
from apps.ledger.choices import InvoiceStatus
from apps.ledger.models import AccountBalanceSnapshot, Invoice, MoneyTransaction

from .choices import BatchStatus, SourceKind
from .fingerprints import transaction_fingerprint
from .models import ImportBatch, SourceFile, StagedRow
from .parsers.agricultural_bank import find_agricultural_bank_sheet
from .registry import ImporterRegistry
from .storage import sha256_file
from .types import (
    DuplicateSourceFileError,
    ImportResult,
    NormalizedInvoiceRow,
    NormalizedTransactionRow,
    ParsedRow,
    RowIssue,
    RowValidationError,
)
from .validation import (
    resolve_funding_account,
    resolve_invoice_counterparty,
    resolve_transaction_counterparty,
    validate_parsed_row,
)

importer_registry = ImporterRegistry()


def stage_upload(file_obj, *, source_kind=None, actor, registry=None) -> ImportBatch:
    require_finance_actor(actor, message="只有财务人员可以执行导入确认")
    registry = registry or importer_registry
    validate_upload_signature(file_obj)
    sha256 = sha256_file(file_obj)
    if SourceFile.objects.filter(sha256=sha256).exists():
        raise DuplicateSourceFileError("相同 SHA-256 的原始文件已导入")

    filename = Path(getattr(file_obj, "name", "upload")).name
    headers = _read_headers(file_obj, filename)
    parser = registry.detect(filename, headers)
    if source_kind is None:
        source_kind = _detect_source_kind(parser, file_obj)
    else:
        source_kind = _validate_source_kind(source_kind)
        if source_kind not in parser.source_kinds:
            raise RowValidationError("解析器不支持该来源类型")

    prepared_rows, counters, period_dates = _prepare_rows(
        parser, file_obj, source_kind
    )
    source_file = None
    try:
        with transaction.atomic():
            batch = ImportBatch.objects.create(source_kind=source_kind, created_by=actor)
            source_file = SourceFile(
                batch=batch,
                file=file_obj,
                original_name=filename,
                sha256=sha256,
                size=_file_size(file_obj),
            )
            try:
                source_file.save()
            except IntegrityError as exc:
                raise DuplicateSourceFileError("相同 SHA-256 的原始文件已导入") from exc
            file_obj.seek(0)
            StagedRow.objects.bulk_create(
                [StagedRow(batch=batch, **prepared_row) for prepared_row in prepared_rows]
            )
            batch.total_rows = counters["total"]
            batch.valid_rows = counters["valid"]
            batch.duplicate_rows = counters["duplicate"]
            batch.error_rows = counters["error"]
            batch.period_start = min(period_dates) if period_dates else None
            batch.period_end = max(period_dates) if period_dates else None
            batch.status = BatchStatus.PREVIEWED
            batch.save(
                update_fields=[
                    "total_rows",
                    "valid_rows",
                    "duplicate_rows",
                    "error_rows",
                    "period_start",
                    "period_end",
                    "status",
                ]
            )
            record_audit(
                actor, "import.staged", batch, ImportResult.from_batch(batch).as_dict()
            )
            return batch
    except Exception:
        _delete_uncommitted_source_file(source_file)
        raise


@transaction.atomic
def confirm_batch(batch_id, actor) -> ImportResult:
    require_finance_actor(actor, message="只有财务人员可以执行导入确认")
    batch = ImportBatch.objects.select_for_update().get(pk=batch_id)
    if batch.confirmed_at is not None:
        return ImportResult.from_batch(batch)

    new_duplicates = 0
    new_errors = 0
    for row in batch.rows.select_for_update().filter(
        posted_at__isnull=True, is_duplicate=False
    ).order_by("row_number"):
        if row.issues:
            continue
        normalized = _deserialize_normalized(row.normalized_data)
        disposition, existing_invoice = _formal_disposition_for(normalized, lock=True)
        if disposition == "duplicate":
            _mark_duplicate(row)
            new_duplicates += 1
            continue
        if disposition == "conflict":
            _mark_invoice_conflict(row, existing_invoice.status, normalized.status)
            new_errors += 1
            continue
        if disposition == "status_update":
            _update_invoice_status(existing_invoice, normalized, batch, row, actor)
            row.posted_at = timezone.now()
            row.save(update_fields=["posted_at"])
            continue
        try:
            with transaction.atomic():
                _post_normalized_row(batch, row)
        except IntegrityError:
            disposition, existing_invoice = _formal_disposition_for(
                normalized, lock=True
            )
            if disposition == "new":
                raise
            if disposition == "duplicate":
                _mark_duplicate(row)
                new_duplicates += 1
                continue
            if disposition == "conflict":
                _mark_invoice_conflict(row, existing_invoice.status, normalized.status)
                new_errors += 1
                continue
            _update_invoice_status(existing_invoice, normalized, batch, row, actor)
        row.posted_at = timezone.now()
        row.save(update_fields=["posted_at"])

    _create_balance_snapshots(batch)

    if new_duplicates:
        batch.duplicate_rows += new_duplicates
        batch.valid_rows -= new_duplicates
    if new_errors:
        batch.error_rows += new_errors
        batch.valid_rows -= new_errors
    batch.confirmed_at = timezone.now()
    batch.status = BatchStatus.PARTIAL if batch.error_rows else BatchStatus.COMPLETED
    batch.save(
        update_fields=[
            "duplicate_rows",
            "error_rows",
            "valid_rows",
            "confirmed_at",
            "status",
        ]
    )
    result = ImportResult.from_batch(batch)
    record_audit(actor, "import.confirmed", batch, result.as_dict())
    return result


def _prepare_rows(parser, file_obj, source_kind):
    prepared_rows = []
    seen_rows = {}
    period_dates = []
    counters = {"total": 0, "valid": 0, "duplicate": 0, "error": 0}
    try:
        file_obj.seek(0)
        for row_count, parsed_row in enumerate(parser.parse(file_obj), start=1):
            if row_count > settings.IMPORT_MAX_ROWS:
                raise RowValidationError("文件数据行数超过系统允许的上限")
            prepared_rows.append(
                _prepare_staged_row(
                    source_kind=source_kind,
                    parsed_row=parsed_row,
                    seen_rows=seen_rows,
                    counters=counters,
                    period_dates=period_dates,
                )
            )
    finally:
        file_obj.seek(0)
    return prepared_rows, counters, period_dates


def _prepare_staged_row(*, source_kind, parsed_row, seen_rows, counters, period_dates):
    if not isinstance(parsed_row, ParsedRow):
        raise RowValidationError("解析器必须返回 ParsedRow")
    counters["total"] += 1
    raw_data, raw_issue = _json_safe_raw_data(parsed_row.raw_data)
    issues = list(validate_parsed_row(source_kind, parsed_row))
    if raw_issue is not None:
        issues.insert(0, raw_issue)
    normalized = parsed_row.normalized
    is_duplicate = False
    if not issues and normalized is not None:
        duplicate_key = _duplicate_key(normalized)
        previous = seen_rows.get(duplicate_key)
        if previous is None:
            disposition, existing_invoice = _formal_disposition_for(normalized)
            current_status = (
                existing_invoice.status if existing_invoice is not None else None
            )
        elif isinstance(normalized, NormalizedInvoiceRow):
            disposition = _invoice_rows_disposition(previous, normalized)
            current_status = previous.status
        else:
            disposition = "duplicate"
            current_status = None
        if disposition != "conflict":
            seen_rows[duplicate_key] = normalized
        if disposition == "duplicate":
            is_duplicate = True
        elif disposition == "conflict":
            issues.append(_invoice_conflict_issue(current_status, normalized.status))
        if isinstance(normalized, NormalizedInvoiceRow):
            period_dates.append(normalized.issue_date)
        else:
            period_dates.append(normalized.occurred_at.date())

    staged_row = {
        "row_number": parsed_row.row_number,
        "raw_data": raw_data,
        "normalized_data": _serialize_normalized(normalized) if not issues else {},
        "issues": [asdict(issue) for issue in issues],
        "is_duplicate": is_duplicate,
    }
    if issues:
        counters["error"] += 1
    elif is_duplicate:
        counters["duplicate"] += 1
    else:
        counters["valid"] += 1
    return staged_row


def _json_safe_raw_data(raw_data):
    try:
        return json.loads(json.dumps(raw_data, cls=DjangoJSONEncoder, ensure_ascii=False)), None
    except (TypeError, ValueError):
        unsupported_value = _first_unserializable_value(raw_data)
        return {
            "_unserializable_type": (
                f"{type(unsupported_value).__module__}."
                f"{type(unsupported_value).__qualname__}"
            ),
            "_unserializable_repr": _safe_repr(raw_data),
        }, RowIssue("raw_data", "原始数据无法序列化", "raw_data")


def _safe_repr(value) -> str:
    return f"<{type(value).__module__}.{type(value).__qualname__}>"


def _first_unserializable_value(value, seen=None):
    seen = seen or set()
    if id(value) in seen:
        return value
    seen.add(id(value))
    if isinstance(value, dict):
        for key, item in value.items():
            if not isinstance(key, (str, int, float, bool, type(None))):
                return key
            unsupported_value = _first_unserializable_value(item, seen)
            if unsupported_value is not None:
                return unsupported_value
        return None
    if isinstance(value, (list, tuple)):
        for item in value:
            unsupported_value = _first_unserializable_value(item, seen)
            if unsupported_value is not None:
                return unsupported_value
        return None
    if isinstance(value, (str, int, float, bool, type(None))):
        return None
    return value


def _delete_uncommitted_source_file(source_file) -> None:
    if source_file is None:
        return
    field_file = source_file.file
    if field_file._committed and field_file.name:
        field_file.storage.delete(field_file.name)


def _post_normalized_row(batch, row) -> None:
    normalized = _deserialize_normalized(row.normalized_data)
    if isinstance(normalized, NormalizedInvoiceRow):
        counterparty = resolve_invoice_counterparty(normalized)
        if counterparty.value is None:
            raise RowValidationError("无法识别往来单位")
        Invoice.objects.create(
            direction=normalized.direction,
            invoice_number=normalized.invoice_number,
            seller_tax_id=normalized.seller_tax_id,
            buyer_tax_id=normalized.buyer_tax_id,
            issue_date=normalized.issue_date,
            total_amount=normalized.total_amount,
            status=normalized.status,
            counterparty=counterparty.value,
            import_batch=batch,
            source_row=row.row_number,
            source_payload=row.raw_data,
        )
        return

    account = resolve_funding_account(normalized)
    counterparty = resolve_transaction_counterparty(normalized)
    if account.value is None:
        raise RowValidationError("无法识别资金账户")
    if counterparty.value is None:
        raise RowValidationError("无法识别往来单位")
    MoneyTransaction.objects.create(
        account=account.value,
        channel=normalized.channel,
        direction=normalized.direction,
        occurred_at=normalized.occurred_at,
        amount=normalized.amount,
        balance_after=normalized.balance_after,
        transaction_id=normalized.transaction_id,
        fingerprint=transaction_fingerprint(normalized),
        counterparty=counterparty.value,
        counterparty_raw_name=normalized.counterparty_name,
        counterparty_account=normalized.counterparty_account,
        import_batch=batch,
        source_row=row.row_number,
        source_payload=row.raw_data,
    )


def _formal_disposition_for(normalized, *, lock=False):
    if isinstance(normalized, NormalizedInvoiceRow):
        invoices = Invoice.objects.filter(
            invoice_number=normalized.invoice_number,
            seller_tax_id=normalized.seller_tax_id,
        )
        if lock:
            invoices = invoices.select_for_update()
        existing = invoices.first()
        if existing is None:
            return "new", None
        return _invoice_disposition(existing, normalized), existing
    is_duplicate = MoneyTransaction.objects.filter(
        fingerprint=transaction_fingerprint(normalized)
    ).exists()
    return ("duplicate" if is_duplicate else "new"), None


def _invoice_disposition(existing: Invoice, normalized: NormalizedInvoiceRow) -> str:
    if not _invoice_business_fields_match(existing, normalized):
        return "conflict"
    if existing.status == normalized.status:
        return "duplicate"
    if existing.status == InvoiceStatus.NORMAL and normalized.status in {
        InvoiceStatus.RED,
        InvoiceStatus.VOID,
    }:
        return "status_update"
    return "conflict"


def _invoice_rows_disposition(
    existing: NormalizedInvoiceRow, normalized: NormalizedInvoiceRow
) -> str:
    if _normalized_invoice_business_fields(existing) != _normalized_invoice_business_fields(
        normalized
    ):
        return "conflict"
    if existing.status == normalized.status:
        return "duplicate"
    if existing.status == InvoiceStatus.NORMAL and normalized.status in {
        InvoiceStatus.RED,
        InvoiceStatus.VOID,
    }:
        return "status_update"
    return "conflict"


def _invoice_business_fields_match(
    existing: Invoice, normalized: NormalizedInvoiceRow
) -> bool:
    persisted_fields = (
        existing.direction,
        existing.invoice_number,
        existing.seller_tax_id,
        existing.buyer_tax_id,
        existing.issue_date,
        existing.total_amount,
    )
    incoming_fields = (
        normalized.direction,
        normalized.invoice_number,
        normalized.seller_tax_id,
        normalized.buyer_tax_id,
        normalized.issue_date,
        Decimal(normalized.total_amount),
    )
    if persisted_fields != incoming_fields:
        return False

    source_data = StagedRow.objects.filter(
        batch_id=existing.import_batch_id,
        row_number=existing.source_row,
    ).values_list("normalized_data", flat=True).first()
    if not source_data or source_data.get("kind") != "invoice":
        return True
    return (
        source_data.get("seller_name") == normalized.seller_name
        and source_data.get("buyer_name") == normalized.buyer_name
    )


def _normalized_invoice_business_fields(row: NormalizedInvoiceRow) -> tuple:
    return (
        row.direction,
        row.invoice_number,
        row.seller_tax_id,
        row.seller_name,
        row.buyer_tax_id,
        row.buyer_name,
        row.issue_date,
        Decimal(row.total_amount),
    )


def _invoice_conflict_issue(current_status, incoming_status) -> RowIssue:
    if current_status in {InvoiceStatus.RED, InvoiceStatus.VOID} and incoming_status == InvoiceStatus.NORMAL:
        return RowIssue(
            "invoice_conflict",
            "红冲或作废发票不允许通过重复导入恢复为正常状态",
        )
    return RowIssue(
        "invoice_conflict",
        "相同发票号码和销售方税号的业务字段冲突",
    )


def _mark_invoice_conflict(row, current_status, incoming_status) -> None:
    row.issues = [asdict(_invoice_conflict_issue(current_status, incoming_status))]
    row.save(update_fields=["issues"])


def _update_invoice_status(existing, normalized, batch, row, actor) -> None:
    previous_status = existing.status
    existing.status = normalized.status
    existing.save(update_fields=["status"])
    record_audit(
        actor,
        "invoice.status_changed",
        existing,
        {
            "from_status": previous_status,
            "to_status": normalized.status,
            "source_batch_id": str(batch.id),
            "source_row": row.row_number,
        },
    )


def _mark_duplicate(row) -> None:
    row.is_duplicate = True
    row.save(update_fields=["is_duplicate"])


def _duplicate_key(normalized):
    if isinstance(normalized, NormalizedInvoiceRow):
        return ("invoice", normalized.invoice_number, normalized.seller_tax_id)
    return ("transaction", transaction_fingerprint(normalized))


def _create_balance_snapshots(batch) -> None:
    if batch.source_kind != SourceKind.BANK:
        return

    latest_by_account = {}
    for row in (
        batch.rows.select_for_update()
        .filter(posted_at__isnull=False)
        .order_by("row_number")
    ):
        if row.issues:
            continue
        normalized = _deserialize_normalized(row.normalized_data)
        if not isinstance(normalized, NormalizedTransactionRow):
            continue
        if normalized.balance_after is None:
            continue
        account = resolve_funding_account(normalized).value
        if account is None:
            continue
        candidate = (normalized.occurred_at, row.row_number, normalized.balance_after)
        previous = latest_by_account.get(account.pk)
        if previous is None or (
            candidate[0] > previous[1]
            or (candidate[0] == previous[1] and candidate[1] < previous[2])
        ):
            latest_by_account[account.pk] = (account, *candidate)

    for account, occurred_at, _row_number, balance in latest_by_account.values():
        _create_balance_snapshot(account, occurred_at, balance, batch)


def _create_balance_snapshot(account, occurred_at, balance, batch) -> None:
    existing = AccountBalanceSnapshot.objects.select_for_update().filter(
        account=account, as_of=occurred_at
    ).first()
    if existing is not None:
        if existing.balance != balance:
            raise RowValidationError("账户余额快照冲突")
        return
    try:
        with transaction.atomic():
            AccountBalanceSnapshot.objects.create(
                account=account,
                as_of=occurred_at,
                balance=balance,
                source_batch=batch,
            )
    except IntegrityError:
        existing = AccountBalanceSnapshot.objects.select_for_update().get(
            account=account, as_of=occurred_at
        )
        if existing.balance != balance:
            raise RowValidationError("账户余额快照冲突")


def _serialize_normalized(normalized):
    if normalized is None:
        return {}
    if isinstance(normalized, NormalizedInvoiceRow):
        return {
            "kind": "invoice",
            "direction": normalized.direction,
            "invoice_number": normalized.invoice_number,
            "seller_tax_id": normalized.seller_tax_id,
            "seller_name": normalized.seller_name,
            "buyer_tax_id": normalized.buyer_tax_id,
            "buyer_name": normalized.buyer_name,
            "issue_date": normalized.issue_date.isoformat(),
            "total_amount": format(Decimal(normalized.total_amount), "f"),
            "status": normalized.status,
        }
    return {
        "kind": "transaction",
        "channel": normalized.channel,
        "direction": normalized.direction,
        "occurred_at": normalized.occurred_at.isoformat(),
        "amount": format(Decimal(normalized.amount), "f"),
        "balance_after": (
            format(Decimal(normalized.balance_after), "f")
            if normalized.balance_after is not None
            else None
        ),
        "account_identifier": normalized.account_identifier,
        "counterparty_name": normalized.counterparty_name,
        "counterparty_account": normalized.counterparty_account,
        "transaction_id": normalized.transaction_id,
        "summary": normalized.summary,
    }


def _deserialize_normalized(data):
    if data.get("kind") == "invoice":
        return NormalizedInvoiceRow(
            direction=data["direction"],
            invoice_number=data["invoice_number"],
            seller_tax_id=data["seller_tax_id"],
            seller_name=data["seller_name"],
            buyer_tax_id=data["buyer_tax_id"],
            buyer_name=data["buyer_name"],
            issue_date=datetime.fromisoformat(data["issue_date"]).date(),
            total_amount=Decimal(data["total_amount"]),
            status=data["status"],
        )
    return NormalizedTransactionRow(
        channel=data["channel"],
        direction=data["direction"],
        occurred_at=datetime.fromisoformat(data["occurred_at"]),
        amount=Decimal(data["amount"]),
        balance_after=(
            Decimal(data["balance_after"]) if data["balance_after"] is not None else None
        ),
        account_identifier=data["account_identifier"],
        counterparty_name=data["counterparty_name"],
        counterparty_account=data["counterparty_account"],
        transaction_id=data["transaction_id"],
        summary=data["summary"],
    )


def _read_headers(file_obj, filename: str) -> list[str]:
    extension = Path(filename).suffix.lower()
    try:
        file_obj.seek(0)
        if extension in {".csv", ".txt"}:
            content = file_obj.read()
            if isinstance(content, str):
                text = content
            else:
                try:
                    text = content.decode("utf-8-sig")
                except UnicodeDecodeError:
                    text = content.decode("gb18030")
            return _densest_row(csv.reader(io.StringIO(text)))
        if extension == ".xlsx":
            workbook = load_workbook(file_obj, read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                return _densest_row(worksheet.iter_rows(values_only=True))
            finally:
                workbook.close()
        if extension == ".xls":
            workbook = xlrd.open_workbook(file_contents=file_obj.read())
            worksheet, _account_identifier = find_agricultural_bank_sheet(workbook)
            return [
                str(value).strip() if value is not None else ""
                for value in worksheet.row_values(2)
            ]
    finally:
        file_obj.seek(0)
    raise RowValidationError("不支持的文件扩展名")


def _densest_row(rows) -> list[str]:
    normalized_rows = [
        [str(value).strip() if value is not None else "" for value in row]
        for row in rows
    ]
    if not normalized_rows:
        return []
    return max(
        enumerate(normalized_rows),
        key=lambda item: (sum(bool(value) for value in item[1]), -item[0]),
    )[1]


def _file_size(file_obj) -> int:
    file_obj.seek(0, io.SEEK_END)
    size = file_obj.tell()
    file_obj.seek(0)
    return size


def _validate_source_kind(source_kind) -> str:
    if source_kind not in SourceKind.values:
        raise ValueError("来源类型不合法")
    return str(source_kind)


def _detect_source_kind(parser, file_obj) -> str:
    source_kinds = tuple(parser.source_kinds)
    if len(source_kinds) == 1:
        return str(source_kinds[0])
    infer_source_kind = getattr(parser, "infer_source_kind", None)
    if infer_source_kind is None:
        raise RowValidationError("无法自动识别文件类型")
    source_kind = _validate_source_kind(infer_source_kind(file_obj))
    if source_kind not in parser.source_kinds:
        raise RowValidationError("解析器无法识别该来源类型")
    return source_kind
