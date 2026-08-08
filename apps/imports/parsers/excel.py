from collections.abc import Iterator

from openpyxl import load_workbook

from apps.imports.types import RowValidationError


def iter_sheet_rows(
    file_obj, sheet_name: str
) -> Iterator[tuple[int, dict[str, object]]]:
    file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RowValidationError(f"Excel 文件缺少 {sheet_name} 工作表")
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(values_only=True), ())
        headers = [
            str(value).strip() if value is not None else "" for value in header_row
        ]
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            yield row_number, dict(zip(headers, values, strict=False))
    finally:
        workbook.close()
        file_obj.seek(0)
