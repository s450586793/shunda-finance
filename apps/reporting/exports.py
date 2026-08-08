from datetime import date, datetime

from django.core.exceptions import PermissionDenied
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from apps.accounts.roles import Role, user_has_role
from apps.core.masking import mask_account
from apps.imports.models import ImportBatch, StagedRow
from apps.ledger.models import Invoice, MoneyTransaction
from apps.reconciliation.models import Reconciliation, ReconciliationAllocation

from .queries import (
    exception_items,
    payables_as_of,
    receivables_as_of,
    unmatched_funds_as_of,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FORMULA_PREFIXES = ("=", "+", "-", "@")


def sanitize_excel_value(value):
    """Prevent external strings from being interpreted as spreadsheet formulas."""
    if isinstance(value, datetime) and timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return f"'{value}"
    return value


def reconciliation_export_rows(reconciliation_ids):
    allocations = (
        ReconciliationAllocation.objects.filter(reconciliation_id__in=reconciliation_ids)
        .select_related(
            "reconciliation",
            "invoice",
            "invoice__counterparty",
            "invoice__import_batch",
            "transaction",
            "transaction__counterparty",
            "transaction__import_batch",
        )
        .prefetch_related(
            "invoice__import_batch__source_files",
            "transaction__import_batch__source_files",
        )
        .order_by("reconciliation_id", "id")
    )
    for allocation in allocations.iterator(chunk_size=500):
        invoice_source_file = next(
            iter(allocation.invoice.import_batch.source_files.all()), None
        )
        transaction_source_file = next(
            iter(allocation.transaction.import_batch.source_files.all()), None
        )
        yield [
            str(allocation.reconciliation_id),
            allocation.reconciliation.created_at,
            allocation.invoice.counterparty.name,
            allocation.transaction.occurred_at,
            allocation.transaction.amount,
            allocation.transaction.channel,
            str(allocation.transaction.import_batch_id),
            allocation.invoice.invoice_number,
            allocation.invoice.total_amount,
            allocation.amount,
            str(transaction_source_file.id) if transaction_source_file else "",
            str(allocation.invoice.import_batch_id),
            str(invoice_source_file.id) if invoice_source_file else "",
        ]


def build_reconciliation_export(reconciliation_ids, actor, *, as_of: date | None = None):
    """Build the finance workbook; full reconciliation detail is finance-only."""
    if not user_has_role(actor, Role.FINANCE):
        raise PermissionDenied("仅财务可以导出完整核销明细")

    as_of = as_of or timezone.localdate()
    workbook = Workbook()
    reconciliation_sheet = workbook.active
    reconciliation_sheet.title = "核销明细"
    _append_sheet(
        reconciliation_sheet,
        [
            "核销ID",
            "核销日期",
            "单位",
            "资金日期",
            "资金金额",
            "渠道",
            "资金来源批次ID",
            "发票号码",
            "发票金额",
            "分配金额",
            "资金来源文件ID",
            "发票来源批次ID",
            "发票来源文件ID",
        ],
        reconciliation_export_rows(reconciliation_ids),
        amount_columns=("E", "I", "J"),
    )
    _append_open_invoice_sheet(workbook, "应收", receivables_as_of(as_of))
    _append_open_invoice_sheet(workbook, "应付", payables_as_of(as_of))
    _append_unmatched_funds_sheet(workbook, as_of)
    _append_exception_sheet(workbook, as_of)
    return workbook


def _append_open_invoice_sheet(workbook, title, rows) -> None:
    sheet = workbook.create_sheet(title)
    _append_sheet(
        sheet,
        [
            "发票ID",
            "开票日期",
            "单位",
            "未核销金额",
            "账龄",
            "到期日",
            "发票来源批次ID",
            "发票来源文件ID",
        ],
        (
            [
                str(row.invoice_id),
                row.issue_date,
                row.counterparty_name,
                row.open_amount,
                row.aging_bucket,
                row.due_date,
                str(row.import_batch_id),
                str(row.source_file_id) if row.source_file_id else "",
            ]
            for row in rows
        ),
        amount_columns=("D",),
    )


def _append_unmatched_funds_sheet(workbook, as_of: date) -> None:
    sheet = workbook.create_sheet("未匹配资金")
    _append_sheet(
        sheet,
        [
            "资金ID",
            "资金日期",
            "单位",
            "未核销金额",
            "渠道",
            "账户",
            "资金来源批次ID",
            "资金来源文件ID",
        ],
        (
            [
                str(row.transaction_id),
                row.occurred_at,
                row.counterparty_name,
                row.open_amount,
                row.channel,
                mask_account(row.masked_account),
                str(row.import_batch_id),
                str(row.source_file_id) if row.source_file_id else "",
            ]
            for row in unmatched_funds_as_of(as_of)
        ),
        amount_columns=("D",),
    )


def _append_exception_sheet(workbook, as_of: date) -> None:
    sheet = workbook.create_sheet("导入异常")
    items = exception_items(as_of)
    source_links = _exception_source_links(items)
    _append_sheet(
        sheet,
        [
            "异常类型",
            "关联ID",
            "日期",
            "单位",
            "金额",
            "说明",
            "来源批次ID",
            "来源文件ID",
        ],
        (
            [
                row.type.value,
                str(row.reference_id),
                row.occurred_on,
                row.counterparty_name,
                row.amount,
                row.detail,
                source_links.get(row.reference_id, ("", ""))[0],
                source_links.get(row.reference_id, ("", ""))[1],
            ]
            for row in items
        ),
        amount_columns=("E",),
    )


def _append_sheet(sheet, headers, rows, *, amount_columns=()) -> None:
    sheet.append([sanitize_excel_value(value) for value in headers])
    for row in rows:
        sheet.append([sanitize_excel_value(value) for value in row])
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for column in amount_columns:
        for cell in sheet[column][1:]:
            cell.number_format = "0.00"
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 36)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)


def _exception_source_links(items):
    reference_ids = {item.reference_id for item in items}
    links = {}
    for batch in ImportBatch.objects.filter(id__in=reference_ids).prefetch_related(
        "source_files"
    ):
        links[batch.id] = _batch_source_link(batch)
    for row in StagedRow.objects.filter(id__in=reference_ids).select_related(
        "batch"
    ).prefetch_related("batch__source_files"):
        links[row.id] = _batch_source_link(row.batch)
    for invoice in Invoice.objects.filter(id__in=reference_ids).select_related(
        "import_batch"
    ).prefetch_related("import_batch__source_files"):
        links[invoice.id] = _batch_source_link(invoice.import_batch)
    for transaction in MoneyTransaction.objects.filter(id__in=reference_ids).select_related(
        "import_batch"
    ).prefetch_related("import_batch__source_files"):
        links[transaction.id] = _batch_source_link(transaction.import_batch)
    reconciliations = Reconciliation.objects.filter(id__in=reference_ids).prefetch_related(
        "allocations__invoice__import_batch__source_files",
        "allocations__transaction__import_batch__source_files",
    )
    for reconciliation in reconciliations:
        batches = []
        for allocation in reconciliation.allocations.all():
            batches.extend((allocation.invoice.import_batch, allocation.transaction.import_batch))
        links[reconciliation.id] = _batches_source_link(batches)
    return links


def _batch_source_link(batch):
    return _batches_source_link([batch])


def _batches_source_link(batches):
    batch_ids = []
    file_ids = []
    for batch in batches:
        if batch.id not in batch_ids:
            batch_ids.append(batch.id)
        for source_file in batch.source_files.all():
            if source_file.id not in file_ids:
                file_ids.append(source_file.id)
    return ";".join(map(str, batch_ids)), ";".join(map(str, file_ids))
